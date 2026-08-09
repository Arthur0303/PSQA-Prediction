from copy import copy
from pathlib import Path

try:
    from openpyxl import load_workbook
    from openpyxl.utils import get_column_letter
except ModuleNotFoundError as exc:
    raise SystemExit(
        "Missing dependency: openpyxl. Install it with: python -m pip install openpyxl"
    ) from exc


SCRIPT_DIR = Path(__file__).resolve().parent
INPUT_FILE = SCRIPT_DIR / "Data" / "training.xlsx"
OUTPUT_FILE = SCRIPT_DIR / "0_remove_columns.xlsx"
SHEET_NAME = "RadCalc"
COLUMNS_TO_REMOVE = (
    "ID",
    "Date",
    "Plan",
    "ALPO(Matlab)",
    "GPR3p2mm",
    "Pass_Fail",
    "Pass",
    "MCS(Matlab)",
)


def is_formula(value: object) -> bool:
    return isinstance(value, str) and value.startswith("=")


def validate_formula_cached_values(
    formula_worksheet,
    values_worksheet,
    retained_column_indices: set[int],
    headers_by_column: dict[int, object],
) -> None:
    missing_cached_values = []

    for row in formula_worksheet.iter_rows(min_row=2):
        for cell in row:
            if cell.column not in retained_column_indices:
                continue
            if not is_formula(cell.value):
                continue

            cached_value = values_worksheet.cell(
                row=cell.row,
                column=cell.column,
            ).value
            if cached_value is None:
                column_letter = get_column_letter(cell.column)
                header = headers_by_column.get(cell.column, "<no header>")
                missing_cached_values.append(
                    f"{column_letter}{cell.row} ({header})"
                )

    if missing_cached_values:
        preview = ", ".join(missing_cached_values[:10])
        if len(missing_cached_values) > 10:
            preview += f", ... and {len(missing_cached_values) - 10} more"
        raise ValueError(
            "Formula cells in retained columns do not have cached values: "
            f"{preview}. Recalculate and save the Excel file before running "
            "this script."
        )


def main() -> None:
    if not INPUT_FILE.is_file():
        raise FileNotFoundError(f"Input file not found: {INPUT_FILE}")

    workbook = load_workbook(INPUT_FILE, data_only=True)
    formula_workbook = load_workbook(INPUT_FILE, data_only=False)
    try:
        if SHEET_NAME not in workbook.sheetnames:
            raise ValueError(f"Worksheet not found: {SHEET_NAME}")

        worksheet = workbook[SHEET_NAME]
        formula_worksheet = formula_workbook[SHEET_NAME]
        columns_by_name = {
            cell.value: cell.column for cell in worksheet[1] if cell.value is not None
        }
        missing_columns = [
            column for column in COLUMNS_TO_REMOVE if column not in columns_by_name
        ]
        if missing_columns:
            missing = ", ".join(missing_columns)
            raise ValueError(f"Columns not found in row 1: {missing}")

        headers_by_column = {cell.column: cell.value for cell in worksheet[1]}
        retained_column_indices = {
            cell.column for cell in worksheet[1] if cell.value not in COLUMNS_TO_REMOVE
        }
        validate_formula_cached_values(
            formula_worksheet,
            worksheet,
            retained_column_indices,
            headers_by_column,
        )

        retained_column_dimensions = []
        for cell in worksheet[1]:
            if cell.value in COLUMNS_TO_REMOVE:
                continue
            dimension = worksheet.column_dimensions[cell.column_letter]
            retained_column_dimensions.append(
                {
                    "width": dimension.width,
                    "hidden": dimension.hidden,
                    "bestFit": dimension.bestFit,
                    "outlineLevel": dimension.outlineLevel,
                    "collapsed": dimension.collapsed,
                }
            )

        column_indices = sorted(
            (columns_by_name[column] for column in COLUMNS_TO_REMOVE),
            reverse=True,
        )
        for column_index in column_indices:
            worksheet.delete_cols(column_index)

        worksheet.column_dimensions.clear()
        for column_index, properties in enumerate(
            retained_column_dimensions,
            start=1,
        ):
            dimension = worksheet.column_dimensions[get_column_letter(column_index)]
            for name, value in properties.items():
                setattr(dimension, name, value)

        for cell in worksheet[1]:
            alignment = copy(cell.alignment)
            alignment.wrap_text = False
            cell.alignment = alignment
        worksheet.row_dimensions[1].height = None

        workbook.save(OUTPUT_FILE)
    finally:
        workbook.close()
        formula_workbook.close()

    print(f"Saved: {OUTPUT_FILE}")


if __name__ == "__main__":
    main()
