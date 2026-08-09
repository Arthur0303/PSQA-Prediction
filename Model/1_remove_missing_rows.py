import csv
from pathlib import Path

try:
    from openpyxl import load_workbook
except ModuleNotFoundError as exc:
    raise SystemExit(
        "Missing dependency: openpyxl. Install it with: "
        "python -m pip install openpyxl"
    ) from exc


SCRIPT_DIR = Path(__file__).resolve().parent
INPUT_FILE = SCRIPT_DIR / "0_remove_columns.xlsx"
OUTPUT_FILE = SCRIPT_DIR / "1_remove_missing_rows.csv"
SHEET_NAME = "RadCalc"


def is_missing(value: object) -> bool:
    return value is None or (isinstance(value, str) and not value.strip())


def main() -> None:
    if not INPUT_FILE.is_file():
        raise FileNotFoundError(f"Input file not found: {INPUT_FILE}")

    workbook = load_workbook(INPUT_FILE, read_only=True, data_only=True)
    try:
        if SHEET_NAME not in workbook.sheetnames:
            raise ValueError(f"Worksheet not found: {SHEET_NAME}")

        worksheet = workbook[SHEET_NAME]
        rows = worksheet.iter_rows(values_only=True)

        try:
            header = next(rows)
        except StopIteration as exc:
            raise ValueError(f"Worksheet is empty: {SHEET_NAME}") from exc

        retained_rows = []
        removed_count = 0

        for row in rows:
            if any(is_missing(value) for value in row):
                removed_count += 1
                continue
            retained_rows.append(row)
    finally:
        workbook.close()

    with OUTPUT_FILE.open("w", encoding="utf-8-sig", newline="") as output:
        writer = csv.writer(output)
        writer.writerow(header)
        writer.writerows(retained_rows)

    print(
        f"Saved: {OUTPUT_FILE} "
        f"(kept {len(retained_rows)} rows, removed {removed_count} rows)"
    )


if __name__ == "__main__":
    main()
