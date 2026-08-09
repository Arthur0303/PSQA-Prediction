"""
Reads a RadCalc PDF with pdfplumber, stores all parameters in a dict, and outputs formatted results.

Python: 3.12.13
Install dependencies:
    pip install pdfplumber openpyxl pydicom numpy
"""

import re
import os
import sys
import math
import pdfplumber
import numpy as np
import pydicom
from pathlib import Path
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

PDF_INITIAL_DIR = Path.home()
DCM_INITIAL_DIR = Path.home()
RTSTRUCT_INITIAL_DIR = Path.home()
HISTORY_FILE_NAME = "history.txt"
HISTORY_DIR_COUNT = 3
PLAN_MONTH_RE = r"JAN|FEB|MAR|APR|MAY|JUN|JUL|AUG|SEP|OCT|NOV|DEC"
PLAN_TOKEN_RE = re.compile(rf"([A-Z0-9]+(?:{PLAN_MONTH_RE})\d{{2}})")
SNC_GPR_MARKERS = (
    "SNC Patient QA of Dose Distribution",
    "Difference (%)",
    "Distance (mm)",
    "Pass (%)",
)
SNC_GPR_COLUMNS_BY_CRITERIA = {
    (3, 3): "GPR3mm",
    (2, 2): "GPR2mm",
    (1, 1): "GPR1mm",
    (3, 2): "GPR3p2mm",
}


def _existing_initial_dir(path: Path) -> str | None:
    return str(path) if path.exists() and path.is_dir() else None


def _blank_history() -> dict[str, list[str]]:
    return {
        "pdf_dirs": [""] * HISTORY_DIR_COUNT,
        "dcm_dirs": [""] * HISTORY_DIR_COUNT,
        "rtstruct_dirs": [""] * HISTORY_DIR_COUNT,
        "messages": [],
    }


def _load_history(history_path: Path) -> dict[str, list[str]]:
    """Load history.txt, returning a blank history if the file is malformed."""
    history = _blank_history()
    try:
        lines = history_path.read_text(encoding="utf-8").splitlines()
    except (OSError, UnicodeDecodeError):
        return history

    if len(lines) < 12 or lines[3].strip() or lines[7].strip() or lines[11].strip():
        return history

    history["pdf_dirs"] = [lines[i].strip() for i in range(3)]
    history["dcm_dirs"] = [lines[i].strip() for i in range(4, 7)]
    history["rtstruct_dirs"] = [lines[i].strip() for i in range(8, 11)]
    history["messages"] = _normalize_history_messages(lines[12:])
    return history


def _normalize_history_messages(messages: list[str]) -> list[str]:
    start = 0
    end = len(messages)
    while start < end and not messages[start].strip():
        start += 1
    while end > start and not messages[end - 1].strip():
        end -= 1

    normalized = []
    previous_blank = False
    for message in messages[start:end]:
        if message.strip():
            normalized.append(message)
            previous_blank = False
        elif not previous_blank:
            normalized.append("")
            previous_blank = True
    return normalized


def _write_history(history_path: Path, history: dict[str, list[str]]) -> None:
    lines = (
        history["pdf_dirs"][:HISTORY_DIR_COUNT]
        + [""]
        + history["dcm_dirs"][:HISTORY_DIR_COUNT]
        + [""]
        + history["rtstruct_dirs"][:HISTORY_DIR_COUNT]
        + [""]
        + _normalize_history_messages(history["messages"])
    )
    history_path.write_text("\n".join(lines) + "\n", encoding="utf-8")


def _save_history(history_path: Path, history: dict[str, list[str]]) -> None:
    try:
        _write_history(history_path, history)
    except OSError as e:
        print(f"Warning: failed to write {history_path.name} | {e}")


def _history_initial_dir(history_dirs: list[str], default_dir: Path) -> Path:
    dirs = history_dirs[:HISTORY_DIR_COUNT]
    if len(dirs) == HISTORY_DIR_COUNT and all(dirs):
        normalized = [os.path.normcase(os.path.normpath(path)) for path in dirs]
        if normalized[0] == normalized[1] == normalized[2]:
            candidate = Path(dirs[0])
            if candidate.exists() and candidate.is_dir():
                return candidate
    return default_dir


def _record_history_dir(history: dict[str, list[str]], key: str, files: list[Path]) -> None:
    if not files:
        return
    history[key] = [str(files[0].parent)] + history[key][: HISTORY_DIR_COUNT - 1]


def _prepend_history_messages(
    history: dict[str, list[str]],
    messages: list[str],
) -> None:
    new_messages = [message for message in messages if message.strip()]
    old_messages = _normalize_history_messages(history["messages"])
    if not new_messages:
        history["messages"] = old_messages
        return
    history["messages"] = new_messages + ([""] + old_messages if old_messages else [])


def _normalize_energy(value: str | None) -> str | None:
    """Normalize RadCalc energy text for spreadsheet output."""
    if value is None:
        return None

    compact = re.sub(r"\s+", "", str(value).strip()).upper()
    m = re.fullmatch(r"(\d+)(?:\.0+)?(MV|F{1,3})?", compact)
    if not m:
        return compact or None

    number, suffix = m.groups()
    return f"{number}{suffix}" if suffix and suffix.startswith("F") else number


def _extract_plan_token(value: str | None) -> str | None:
    """Extract a stable plan token such as PELVISMAR26 from RadCalc/TPS names."""
    if value is None:
        return None

    compact = re.sub(r"\s+", "", str(value).strip()).upper()
    m = PLAN_TOKEN_RE.search(compact)
    return m.group(1) if m else None


def _extract_snc_labeled_number(text: str, label: str) -> float | None:
    pattern = rf"(?im)^\s*{re.escape(label)}\s*:?\s*(-?\d+(?:\.\d+)?)\b"
    m = re.search(pattern, text)
    return float(m.group(1)) if m else None


def _snc_criterion_value(value: float | None) -> int | None:
    if value is None:
        return None
    rounded = round(value)
    return int(rounded) if math.isclose(value, rounded, abs_tol=1e-9) else None


def _parse_snc_gpr_from_text(text: str) -> dict[str, float]:
    """Extract SNC gamma pass rate values from one PDF page of text."""
    if not all(marker in text for marker in SNC_GPR_MARKERS):
        return {}

    difference = _snc_criterion_value(
        _extract_snc_labeled_number(text, "Difference (%)")
    )
    distance = _snc_criterion_value(
        _extract_snc_labeled_number(text, "Distance (mm)")
    )
    pass_value = _extract_snc_labeled_number(text, "Pass (%)")
    if difference is None or distance is None or pass_value is None:
        return {}

    column = SNC_GPR_COLUMNS_BY_CRITERIA.get((difference, distance))
    return {column: pass_value} if column else {}


def extract_snc_gpr_values(pdf_path: str) -> dict[str, float]:
    """Extract supported SNC GPR values from all pages of a PDF."""
    gpr_values = {}
    with pdfplumber.open(pdf_path) as pdf:
        for page in pdf.pages:
            gpr_values.update(_parse_snc_gpr_from_text(page.extract_text() or ""))
    return gpr_values


def extract_radcalc_params(pdf_path: str) -> dict:
    """Extract all parameters from a RadCalc PDF and return a nested dict."""

    # Locate the page that matches the expected RadCalc layout. The target
    # page is not guaranteed to be page 1, but its format is consistent —
    # identify it by distinctive markers that co-occur on that page.
    markers = ("Patient Name:", "Calculation Name:", "Beam Description")

    page_text = None
    target_page_index = None
    snc_gpr_values = {}
    with pdfplumber.open(pdf_path) as pdf:
        for page_index, page in enumerate(pdf.pages):
            text = page.extract_text() or ""
            snc_gpr_values.update(_parse_snc_gpr_from_text(text))
            if page_text is None and all(mk in text for mk in markers):
                page_text = text
                target_page_index = page_index

    if page_text is None:
        raise ValueError(
            f"No page matching the expected RadCalc layout found in {pdf_path}"
        )

    lines = page_text.split("\n")
    full_text = page_text  # for regex searches across lines

    already_deidentified = Path(pdf_path).stem.upper().endswith("_MOCK")
    mock_id_from_filename = re.sub(r"_MOCK$", "", Path(pdf_path).stem, flags=re.IGNORECASE)

    data = {
        "block1": {},  # Patient info
        "block2": {},  # Prescription info
        "block3": {},  # Beam parameters
        "block4": {},  # Sign-off
        "gpr": snc_gpr_values,
        "_already_deidentified": already_deidentified,
        "_warnings": [],
    }

    # ── Helper: search a value by regex in the full text ──
    def find(pattern, text=full_text, group=1):
        m = re.search(pattern, text)
        return m.group(group).strip() if m else None

    def find_all(pattern, text=full_text, group=1):
        """Like ``find`` but returns every match (one per arc, in arc order)."""
        return [m.group(group).strip() for m in re.finditer(pattern, text)]

    # ── Helper: try to cast to float, fallback to string ──
    def num(val):
        if val is None:
            return None
        val = val.strip().rstrip("%")
        try:
            return int(val) if "." not in val else float(val)
        except ValueError:
            return val

    # ══════════════════════════════════════════════════════════
    # Block 1 – Patient Info
    # ══════════════════════════════════════════════════════════
    patient_line = next((line for line in lines if "Patient Name:" in line and "Patient ID#:" in line), "")
    m = re.search(
        r"Patient Name:[ \t]*(.*?)[ \t]*,?[ \t]*Patient ID#:[ \t]*([A-Za-z0-9._-]+)\b",
        patient_line,
    )
    if m:
        patient_id = m.group(2).strip()
        if patient_id.upper() not in {"CALCULATION", "COMMENTS", "PRESCRIPTION"}:
            data["block1"]["Patient Name"] = m.group(1).strip().rstrip(",")
            data["block1"]["Patient ID"] = patient_id
    if already_deidentified and mock_id_from_filename:
        data["block1"]["Patient ID"] = mock_id_from_filename
    if "Patient ID" not in data["block1"]:
        raise ValueError(f"Cannot find a valid Patient ID on the patient info line in {pdf_path}")

    raw_calc_name = find(r"Calculation Name:\s*([^\r\n]+)")
    data["_raw_calculation_name"] = raw_calc_name
    data["block1"]["Calculation Name"] = _extract_plan_token(raw_calc_name) or raw_calc_name

    # Prescription Name is the first token on the line after the header row
    for i, line in enumerate(lines):
        if line.startswith("Prescription Name"):
            if i + 1 < len(lines):
                parts = lines[i + 1].split()
                if parts:
                    data["block1"]["Prescription Name"] = parts[0]
            break

    # ══════════════════════════════════════════════════════════
    # Block 2 – Prescription
    # ══════════════════════════════════════════════════════════
    for i, line in enumerate(lines):
        if line.startswith("Prescription Name"):
            if i + 1 < len(lines):
                parts = lines[i + 1].split()
                # e.g. ['ENDOMAR26', '45.000', '1.800', '25.0']
                if len(parts) >= 4:
                    data["block2"]["Prescribed Dose (Gy)"] = num(parts[1])
                    data["block2"]["Dose Per Treatment (Gy)"] = num(parts[2])
                    data["block2"]["Num Fractions"] = num(parts[3])
            break

    # ══════════════════════════════════════════════════════════
    # Block 3 – Beam Parameters
    # ══════════════════════════════════════════════════════════
    b3 = data["block3"]

    # --- Beam Description (line after "Beam Description") ---
    for i, line in enumerate(lines):
        if line.startswith("Beam Description"):
            if i + 1 < len(lines):
                # Next line: e.g. "ARC1-2 PELVIS Energy:6.0 MV"
                desc_line = lines[i + 1]
                # Take everything before "Energy:" as the description
                desc_part = re.split(r"\s+Energy:", desc_line)[0].strip()
                b3["Beam Description"] = desc_part
            break

    # --- Simple key: value patterns on each line (left column) ---
    simple_left = {
        r"Beam ID:\s*(\S+)":                "Beam ID",
        r"Calc Pt:\s*(\S+)":                "Calc Pt",
        r"Dose Per Treat\.:\s*([\d.]+)\s*Gy": "Dose Per Treat.",
        r"Iso\. @ Calc Pt:\s*([\d.]+)%":    "Iso. @ Calc Pt",
        r"RTP Dose @ Calc:\s*([\d.]+)\s*Gy": "RTP Dose @ Calc",
        r"Wedge Factor:\s*([\d.]+)":         "Wedge Factor",
        r"Sc:\s*([\d.]+)":                   "Sc",
        r"ISF:\s*([\d.]+)":                  "ISF",
        r"Primary x OAF:\s*([\d.]+)":        "Primary x OAF",
        r"Scatter:\s*([\d.]+)":              "Scatter",
        r"Inhomo\. Corr\.:\s*([\d.]+)":      "Inhomo. Corr.",
        r"Calib\. at Ref\.:\s*([\d.]+)\s*Gy": "Calib. at Ref.",
        r"Dose Per MU:\s*([\d.]+)\s*":       "Dose Per MU",
        r"(?<!Per )(?<!Plan )MU:\s*([\d.]+)": "MU",
        r"Plan MU:\s*([\d.]+)":              "Plan MU",
        r"Percent Diff:\s*(-?[\d.]+)%":      "Percent Diff",
    }

    # Fields that carry a unit suffix – keep the raw string to preserve
    # original decimal precision (e.g. "0.01000" instead of 0.01)
    unit_fields = {
        "Dose Per Treat.": "Gy",
        "Iso. @ Calc Pt": "%",
        "RTP Dose @ Calc": "Gy",
        "Calib. at Ref.": "Gy",
        "Dose Per MU": "Gy",
        "Percent Diff": "%",
    }

    for pattern, key in simple_left.items():
        val = find(pattern)
        if val is not None:
            if key in unit_fields:
                b3[key] = f"{val} {unit_fields[key]}"
            else:
                b3[key] = num(val)

    # Some RadCalc PDFs have a text-flow order that breaks the MU / percent
    # difference rows into single-character lines. A tighter y tolerance
    # preserves those rows, but it can damage other fields, so only use it to
    # fill missing values known to benefit from it.
    fallback_patterns = {
        "Dose Per MU": (r"Dose Per MU:\s*([\d.]+)\s*", "Gy"),
        "MU": (r"(?<!Per )(?<!Plan )MU:\s*([\d.]+)", None),
        "Plan MU": (r"Plan MU:\s*([\d.]+)", None),
        "Percent Diff": (r"Percent Diff:\s*(-?[\d.]+)\s*%", "%"),
    }
    missing_fallback_fields = {
        key: pattern_unit
        for key, pattern_unit in fallback_patterns.items()
        if b3.get(key) is None
    }
    if missing_fallback_fields and target_page_index is not None:
        with pdfplumber.open(pdf_path) as pdf:
            fallback_text = pdf.pages[target_page_index].extract_text(y_tolerance=1) or ""

        for key, (pattern, unit) in missing_fallback_fields.items():
            val = find(pattern, text=fallback_text)
            if val is not None:
                b3[key] = f"{val} {unit}" if unit else num(val)

    # --- Right-column parameters (often on same line, colon-separated) ---
    b3["Accelerator"] = find(r"Accelerator:(\S+)")
    energy = find(r"Energy:\s*((?i:\d+(?:\.\d+)?\s*(?:MV|F{1,3})?))\b")
    if energy:
        b3["Energy"] = _normalize_energy(energy)

    b3["Gantry"] = find(r"Gantry:(.*?)(?:\n|$)", full_text)
    if b3.get("Gantry"):
        # Clean: may have trailing text from next key
        b3["Gantry"] = re.split(r"\s{2,}", b3["Gantry"])[0].strip()

    collim = find(r"Collim\.:\s*(\S+)")
    if collim:
        b3["Collim"] = num(collim)

    couch = find(r"Couch:\s*(\S+)")
    if couch:
        b3["Couch"] = num(couch)

    ssd = find(r"SSD:([\d.]+)")
    if ssd:
        b3["SSD"] = num(ssd)

    avg_depth = find(r"Avg\. Depth:([\d.]+)")
    if avg_depth:
        b3["Avg. Depth"] = num(avg_depth)

    ssd_iso = find(r"SSD @ Iso:([\d.]+)")
    if ssd_iso:
        b3["SSD @ Iso"] = num(ssd_iso)

    avg_eff = find(r"Avg\. Eff\. Depth:([\d.]+)")
    if avg_eff:
        b3["Avg. Eff. Depth"] = num(avg_eff)

    eq_sq = find(r"Eq\. Sq\.:([\d.]+)")
    if eq_sq:
        b3["Eq. Sq."] = num(eq_sq)

    alpo = find(r"ALPO:([\d.]+)")
    if alpo:
        b3["ALPO"] = num(alpo)

    ncp = find(r"Number of Control Points:\s*(\d+)")
    if ncp:
        b3["Number of Control Points"] = int(ncp)

    mod_factor = find(r"Mod Factor\s*=\s*([\d.]+)")
    if mod_factor:
        b3["Mod Factor"] = num(mod_factor)

    # ══════════════════════════════════════════════════════════
    # Multi-arc aggregation (RadCalc fields)
    # ══════════════════════════════════════════════════════════
    # RadCalc lays each arc out as its own side-by-side beam block; pdfplumber
    # reads them left->right per row, so re.findall returns one value per arc in
    # arc order. The single-value parsing above (re.search) only captured ARC1.
    # For >1 arc, replace the affected fields with the correct plan-level value:
    #   Plan MU / CP -> sum;  the rest -> MU-weighted average (weight = arc Plan MU).
    # PMU is derived downstream from the (now summed) Plan MU. Single-arc plans
    # have one match, so this block is skipped and their output is unchanged.
    # Arc count comes from the per-block "Beam Description" header (one per arc),
    # which sits above the intensity-map graphic and is never shredded.
    n_arcs = full_text.count("Beam Description")
    if n_arcs > 1:
        pdf_name = Path(pdf_path).name
        # Per-arc rows can be broken up by the intensity-map overlay, and which
        # rows break differs between the default and tighter-y-tolerance text
        # extractions. Keep both and let each field use whichever one yields one
        # clean value per arc (values read left->right, i.e. in arc order).
        texts = [full_text]
        if target_page_index is not None:
            with pdfplumber.open(pdf_path) as pdf:
                yt_text = pdf.pages[target_page_index].extract_text(y_tolerance=1) or ""
            texts.append(yt_text)

        def arc_values(pattern):
            """Per-arc match strings from whichever text gives exactly n_arcs."""
            for txt in texts:
                ms = find_all(pattern, text=txt)
                if len(ms) == n_arcs:
                    return ms
            return None

        plan_mu_strs = arc_values(r"Plan MU:\s*([\d.]+)")
        if plan_mu_strs is None:
            data["_warnings"].append(
                f"Warning: {pdf_name} | {n_arcs} arcs but Plan MU not recovered per arc; "
                f"RadCalc fields left as first arc"
            )
        else:
            plan_mu = [num(v) for v in plan_mu_strs]
            total_mu = sum(plan_mu)
            weights = (
                [m / total_mu for m in plan_mu] if total_mu else [1.0 / n_arcs] * n_arcs
            )

            # Sum fields
            b3["Plan MU"] = round(total_mu, 2)
            cp_strs = arc_values(r"Number of Control Points:\s*(\d+)")
            if cp_strs is not None:
                b3["Number of Control Points"] = sum(int(v) for v in cp_strs)
            else:
                data["_warnings"].append(
                    f"Warning: {pdf_name} | {n_arcs} arcs but CP not recovered per arc; "
                    f"CP left as first arc"
                )

            # MU-weighted-average fields (round to each field's RadCalc precision;
            # Percent Diff keeps the "<val> %" string form so _num_only still parses it)
            weighted_specs = [
                ("Percent Diff",    r"Percent Diff:\s*(-?[\d.]+)%", 2, "%"),
                ("Mod Factor",      r"Mod Factor\s*=\s*([\d.]+)",   3, None),
                ("ALPO",            r"ALPO:([\d.]+)",               2, None),
                ("Eq. Sq.",         r"Eq\. Sq\.:([\d.]+)",          2, None),
                ("Avg. Depth",      r"Avg\. Depth:([\d.]+)",        2, None),
                ("Avg. Eff. Depth", r"Avg\. Eff\. Depth:([\d.]+)",  2, None),
            ]
            for key, pattern, ndigits, unit in weighted_specs:
                strs = arc_values(pattern)
                vals = [num(v) for v in strs] if strs is not None else None
                if vals is None or not all(isinstance(v, (int, float)) for v in vals):
                    data["_warnings"].append(
                        f"Warning: {pdf_name} | {n_arcs} arcs but inconsistent '{key}' "
                        f"values; left as first arc"
                    )
                    continue
                wavg = round(sum(v * w for v, w in zip(vals, weights)), ndigits)
                b3[key] = f"{wavg} {unit}" if unit else wavg

    # ══════════════════════════════════════════════════════════
    # Block 4 – Sign-off
    # ══════════════════════════════════════════════════════════
    calc_by = re.search(r"Calculated By:\s*(.+?)\s+(\d{2}/\d{2}/\d{4})", full_text)
    if calc_by:
        data["block4"]["Calculated By"] = f"{calc_by.group(1).strip()}  {calc_by.group(2)}"

    check_by = re.search(r"Checked By:\s*(.+?)\s+(\d{2}/\d{2}/\d{4})", full_text)
    if check_by:
        data["block4"]["Checked By"] = f"{check_by.group(1).strip()}  {check_by.group(2)}"

    return data


def print_radcalc(data: dict, file_name: str = "") -> None:
    """Pretty-print the extracted RadCalc parameters."""

    CYAN = "\033[96m"
    GREEN = "\033[92m"
    YELLOW = "\033[93m"
    RESET = "\033[0m"
    BOLD = "\033[1m"

    width = 64
    sep = "─" * width

    if file_name:
        print(f"\n{BOLD}{CYAN}{'═' * width}{RESET}")
        print(f"{BOLD}{CYAN}  {file_name}{RESET}")
        print(f"{BOLD}{CYAN}{'═' * width}{RESET}")

    block_labels = {
        "block1": "Patient Info",
        "block2": "Prescription",
        "block3": "Beam Parameters",
        "block4": "Sign-off",
    }

    for block_key, label in block_labels.items():
        block = data.get(block_key, {})
        if not block:
            continue
        print(f"\n{YELLOW}  [{label}]{RESET}")
        print(f"  {sep}")
        for key, val in block.items():
            print(f"  {GREEN}{key:<30}{RESET} {val}")

    print()


def _extract_site(calc_name: str) -> str:
    """Return everything before the first month abbreviation in Calculation Name.
    e.g. 'ENDOMAR26' → 'ENDO',  'PROSBEDMAR26' → 'PROSBED'
    """
    calc_upper = calc_name.upper()
    m = re.search(r"(JAN|FEB|MAR|APR|MAY|JUN|JUL|AUG|SEP|OCT|NOV|DEC)", calc_upper)
    return calc_upper[:m.start()] if m else calc_name


def _extract_date(block4_calc_by: str) -> str:
    """Pull the dd/mm/yyyy date from the 'Calculated By' value."""
    m = re.search(r"(\d{2}/\d{2}/\d{4})", block4_calc_by or "")
    return m.group(1) if m else ""


def _num_only(val) -> object:
    """Return the numeric portion of a string like '6.0 MV' or '-2.1 %'."""
    if val is None:
        return None
    s = str(val).strip()
    m = re.match(r"(-?[\d.]+)", s)
    if m:
        v = m.group(1)
        return int(v) if "." not in v else float(v)
    return None


def _round_metrics(metrics: dict, ndigits: int = 4) -> dict:
    """Round numeric metric values for spreadsheet output."""
    return {
        k: (v if isinstance(v, (int, np.integer)) and not isinstance(v, bool)
            else round(float(v), ndigits)
            if isinstance(v, (float, np.floating)) and math.isfinite(float(v))
            else v)
        for k, v in metrics.items()
    }


def _circular_angle_delta(start: float, end: float) -> float:
    """Return the smallest absolute angle change between two 0-360 positions."""
    return abs((end - start + 180.0) % 360.0 - 180.0)


def _gan_from_angles(angles: list[float]) -> float | None:
    """Total accumulated gantry rotation over the full control-point sequence."""
    if len(angles) < 2:
        return None
    total = sum(
        _circular_angle_delta(a, b)
        for a, b in zip(angles, angles[1:])
    )
    return int(round(total)) if math.isclose(total, round(total), abs_tol=1e-6) else round(total, 4)


def extract_gan_from_plan(plan_path: str) -> float | None:
    """Extract accumulated gantry rotation from a DICOM RTPLAN."""
    path = Path(plan_path)

    ds = pydicom.dcmread(str(path), force=True, stop_before_pixels=True)
    if getattr(ds, "Modality", None) != "RTPLAN":
        raise ValueError(f"Not an RTPLAN: {plan_path}")

    totals = []
    for beam in getattr(ds, "BeamSequence", []):
        angles = []
        for cp in getattr(beam, "ControlPointSequence", []):
            angle = getattr(cp, "GantryAngle", None)
            if angle is not None:
                angles.append(float(angle))
        gan = _gan_from_angles(angles)
        if gan is not None:
            totals.append(gan)
    return sum(totals) if totals else None


def _aperture_perimeter(xA: np.ndarray, xB: np.ndarray, h: np.ndarray) -> float:
    """Perimeter of the MLC aperture: vertical leaf edges where open, horizontal
    step edges between adjacent leaves, plus top/bottom edges of the open region."""
    gap = np.maximum(0.0, xB - xA)
    open_mask = gap > 0
    if not np.any(open_mask):
        return 0.0
    P = 2.0 * float(np.sum(h[open_mask]))  # left + right vertical edges
    L = len(h)
    for j in range(L - 1):
        a, b = open_mask[j], open_mask[j + 1]
        if a and b:
            # The exposed boundary between two leaf rows is the symmetric
            # difference of their open intervals.  The usual endpoint-difference
            # shortcut works only when those intervals overlap; for disjoint
            # intervals it incorrectly counts the empty space between them.
            overlap = max(
                0.0,
                min(float(xB[j]), float(xB[j + 1]))
                - max(float(xA[j]), float(xA[j + 1])),
            )
            P += float(gap[j] + gap[j + 1] - 2.0 * overlap)
        elif a and not b:
            P += float(gap[j])
        elif b and not a:
            P += float(gap[j + 1])
    # Array-end caps: leaf 0's top edge and the last leaf's bottom edge are exposed
    # only if those end leaves are themselves open. Interior block tops/bottoms are
    # already counted by the open<->closed transitions above; using the first/last
    # *open* leaf here would double-count them.
    if open_mask[0]:
        P += float(gap[0])
    if open_mask[-1]:
        P += float(gap[-1])
    return P


def _mean_leaf_travel(cps: list[dict], movement_tol: float = 1e-9) -> float:
    """Return Masi LT for one arc from raw mechanical leaf positions.

    A bank leaf is included only when its opposing pair opens inside the jaws at
    least once and that individual leaf moves during the arc.  This excludes
    permanently closed and stationary leaves as required by Masi 2013.  When no
    leaf is eligible (a static MLC), LT is zero.
    """
    if len(cps) < 2:
        return 0.0

    ever_open = np.zeros(len(cps[0]["gap"]), dtype=bool)
    for cp in cps:
        ever_open |= (cp["h_eff"] > 0) & (cp["gap"] > 0)

    xA = np.stack([cp["xA_raw"] for cp in cps])
    xB = np.stack([cp["xB_raw"] for cp in cps])
    travel_A = np.sum(np.abs(np.diff(xA, axis=0)), axis=0)
    travel_B = np.sum(np.abs(np.diff(xB, axis=0)), axis=0)

    include_A = ever_open & (travel_A > movement_tol)
    include_B = ever_open & (travel_B > movement_tol)
    leaf_count = int(np.count_nonzero(include_A) + np.count_nonzero(include_B))
    if leaf_count == 0:
        return 0.0

    total_travel = float(np.sum(travel_A[include_A]) + np.sum(travel_B[include_B]))
    return total_travel / leaf_count


def _jaw_extent(positions: dict, mlc_type: str):
    """Return (travel_min, travel_max, cross_min, cross_max) for the given MLC
    orientation. ASYMX clips along the X axis, ASYMY along Y. MLCX leaves travel
    along X (cross-leaf axis is Y); MLCY is the other way around. Missing jaws
    fall back to wide-open."""
    if mlc_type == "MLCX":
        travel = positions.get("ASYMX") or positions.get("X")
        cross  = positions.get("ASYMY") or positions.get("Y")
    else:  # MLCY
        travel = positions.get("ASYMY") or positions.get("Y")
        cross  = positions.get("ASYMX") or positions.get("X")
    tmin, tmax = (min(travel), max(travel)) if travel else (-1e9, 1e9)
    cmin, cmax = (min(cross),  max(cross))  if cross  else (-1e9, 1e9)
    return tmin, tmax, cmin, cmax


def _clip_aperture(xA: np.ndarray, xB: np.ndarray, boundaries: np.ndarray,
                   jaws: tuple) -> tuple:
    """Clip leaf positions to jaw extents. Returns (xA', xB', h_eff) where
    h_eff is the effective leaf height after Y-jaw clipping. Leaves with no
    Y-overlap are forced closed so they contribute nothing to area/perimeter."""
    tmin, tmax, cmin, cmax = jaws
    xA_c = np.minimum(np.maximum(xA, tmin), tmax)
    xB_c = np.minimum(np.maximum(xB, tmin), tmax)
    b_low, b_high = boundaries[:-1], boundaries[1:]
    h_eff = np.maximum(0.0, np.minimum(b_high, cmax) - np.maximum(b_low, cmin))
    closed = h_eff <= 0
    if np.any(closed):
        # collapse closed leaves to a single point so gap=0
        mid = 0.5 * (xA_c + xB_c)
        xA_c = np.where(closed, mid, xA_c)
        xB_c = np.where(closed, mid, xB_c)
    return xA_c, xB_c, h_eff


def _lsv_bank(x: np.ndarray) -> float:
    """LSV for a single MLC bank (Masi 2013): 1 means uniform, < 1 means irregular."""
    L = len(x)
    if L < 2:
        return 1.0
    pmax = float(np.max(x) - np.min(x))
    if pmax <= 0:
        return 1.0
    diffs = np.abs(np.diff(x))
    return float(np.sum(pmax - diffs) / ((L - 1) * pmax))


def _matlab_lsv_bank(x: np.ndarray) -> float | None:
    """LSV bank calculation copied from MCS_only_7Mar24_1AM.m."""
    L = len(x)
    if L < 2:
        return None
    pmax = float(np.max(x) - np.min(x))
    if pmax == 0:
        return None
    diffs = np.abs(np.diff(x))
    return float(np.sum(pmax - diffs) / ((L - 1) * pmax))


def _matlab_cumulative_dose_coefficient(cp) -> float | None:
    refs = getattr(cp, "ReferencedDoseReferenceSequence", None)
    if not refs:
        return None
    value = getattr(refs[0], "CumulativeDoseReferenceCoefficient", None)
    if value is None:
        return None
    try:
        coeff = float(value)
    except (TypeError, ValueError):
        return None
    return coeff if math.isfinite(coeff) else None


def _calculate_mcs_matlab(ds) -> float | None:
    """Calculate MCS using the formula in MCS_only_7Mar24_1AM.m.

    The MATLAB script leaves ``fl`` at the last BeamSequence item before
    calculating MCS, so this intentionally uses only the last beam.
    """
    beams = list(getattr(ds, "BeamSequence", []))
    if not beams:
        return None

    beam = beams[-1]
    cps = list(getattr(beam, "ControlPointSequence", []))
    if len(cps) < 2:
        return None

    leaf_y = 197.5 - 5.0 * np.arange(80, dtype=float)
    cum_dose = []
    lsv_seg = []
    aav_seg = []

    for cp in cps:
        coeff = _matlab_cumulative_dose_coefficient(cp)
        position_seq = getattr(cp, "BeamLimitingDevicePositionSequence", None)
        if coeff is None or position_seq is None or len(position_seq) < 2:
            return None

        jaw_positions = getattr(position_seq[0], "LeafJawPositions", None)
        mlc_positions = getattr(position_seq[1], "LeafJawPositions", None)
        if jaw_positions is None or mlc_positions is None:
            return None

        try:
            jaw = np.asarray(jaw_positions, dtype=float)
            mlc = np.asarray(mlc_positions, dtype=float)
        except (TypeError, ValueError):
            return None

        if len(jaw) < 2 or len(mlc) != 160:
            return None

        mlc_left = mlc[:80][::-1]
        mlc_right = mlc[80:][::-1]
        active = (leaf_y <= jaw[1]) & (leaf_y >= jaw[0])
        inner_left = mlc_left[active]
        inner_right = mlc_right[active]
        leaf_count = len(inner_left)
        if leaf_count < 2:
            return None

        lsv_left = _matlab_lsv_bank(inner_left)
        lsv_right = _matlab_lsv_bank(inner_right)
        aav_den = leaf_count * (float(np.min(inner_left)) - float(np.max(inner_right)))
        if lsv_left is None or lsv_right is None or aav_den == 0:
            return None

        lsv = lsv_left * lsv_right
        aav = float(np.sum(inner_left - inner_right)) / aav_den
        if not math.isfinite(lsv) or not math.isfinite(aav):
            return None

        cum_dose.append(coeff)
        lsv_seg.append(lsv)
        aav_seg.append(aav)

    div_dose = np.diff(np.asarray(cum_dose, dtype=float))
    lsv_avg = (np.asarray(lsv_seg[:-1], dtype=float) + np.asarray(lsv_seg[1:], dtype=float)) / 2.0
    aav_avg = (np.asarray(aav_seg[:-1], dtype=float) + np.asarray(aav_seg[1:], dtype=float)) / 2.0
    mcs = float(np.sum(lsv_avg * aav_avg * div_dose))
    return mcs if math.isfinite(mcs) else None


def _calculate_alpo_matlab(ds) -> float | None:
    """Calculate ALPO using the formula in MCS_only_7Mar24_1AM.m.

    This mirrors the MATLAB script's last-beam behavior and converts mm to cm.
    """
    beams = list(getattr(ds, "BeamSequence", []))
    if not beams:
        return None

    beam = beams[-1]
    cps = list(getattr(beam, "ControlPointSequence", []))
    if len(cps) < 2:
        return None

    leaf_y = 197.5 - 5.0 * np.arange(80, dtype=float)
    cum_dose = []
    xrt_ltsum = []

    for cp in cps:
        coeff = _matlab_cumulative_dose_coefficient(cp)
        position_seq = getattr(cp, "BeamLimitingDevicePositionSequence", None)
        if coeff is None or position_seq is None or len(position_seq) < 2:
            return None

        jaw_positions = getattr(position_seq[0], "LeafJawPositions", None)
        mlc_positions = getattr(position_seq[1], "LeafJawPositions", None)
        if jaw_positions is None or mlc_positions is None:
            return None

        try:
            jaw = np.asarray(jaw_positions, dtype=float)
            mlc = np.asarray(mlc_positions, dtype=float)
        except (TypeError, ValueError):
            return None

        if len(jaw) < 2 or len(mlc) != 160:
            return None

        mlc_left = mlc[:80][::-1]
        mlc_right = mlc[80:][::-1]
        active = (leaf_y <= jaw[1]) & (leaf_y >= jaw[0])
        inner_left = mlc_left[active]
        inner_right = mlc_right[active]
        if len(inner_left) == 0:
            return None

        cp_alpo = float(np.mean(np.abs(inner_right - inner_left)))
        if not math.isfinite(cp_alpo):
            return None

        cum_dose.append(coeff)
        xrt_ltsum.append(cp_alpo)

    div_dose = np.diff(np.asarray(cum_dose, dtype=float))
    xrt_ltsum_avg = (
        np.asarray(xrt_ltsum[:-1], dtype=float) + np.asarray(xrt_ltsum[1:], dtype=float)
    ) / 2.0
    alpo = float(np.sum(xrt_ltsum_avg * div_dose) / 10.0)
    return alpo if math.isfinite(alpo) else None


def _device_positions(cp, previous: dict) -> dict:
    """Return CP device positions, carrying forward DICOM values omitted on
    this CP (DICOM only stores changed positions after the first CP)."""
    positions = dict(previous)
    for d in getattr(cp, "BeamLimitingDevicePositionSequence", []):
        dev_type = str(getattr(d, "RTBeamLimitingDeviceType", "")).upper()
        positions[dev_type] = [float(x) for x in d.LeafJawPositions]
    return positions


def _arccheck_from_y_jaws(ds) -> int | None:
    """Return ArcCheck setting from max Y jaw opening across all control points.

    RTPLAN jaw positions are stored in mm. The ArcCheck threshold is evaluated
    in cm as max(Y1 opening) + max(Y2 opening), matching the MATLAB Y1/Y2 logic.
    """
    max_y1_cm = None
    max_y2_cm = None

    for beam in getattr(ds, "BeamSequence", []):
        positions = {}
        for cp in getattr(beam, "ControlPointSequence", []):
            positions = _device_positions(cp, positions)
            y_jaws = positions.get("ASYMY") or positions.get("Y")
            if not y_jaws or len(y_jaws) < 2:
                continue

            y1_cm = abs(float(y_jaws[0])) / 10.0
            y2_cm = abs(float(y_jaws[1])) / 10.0
            max_y1_cm = y1_cm if max_y1_cm is None else max(max_y1_cm, y1_cm)
            max_y2_cm = y2_cm if max_y2_cm is None else max(max_y2_cm, y2_cm)

    if max_y1_cm is None or max_y2_cm is None:
        return None
    return 2 if max_y1_cm + max_y2_cm >= 20.5 else 1


def extract_complexity_metrics(rtplan_path: str) -> dict:
    """Compute VMAT complexity metrics from an RTPLAN DICOM file.
    Distances are mm; areas are mm^2.

    Per-beam metrics follow Masi 2013 (AAV normalized to the sum of each leaf
    pair's maximum gap over the arc, CP-MU weighted within beam). Plan-level
    metrics are MU-weighted across beams. Apertures are clipped to ASYMX/ASYMY
    jaws. AAV/LSV/SAS treat any positive cross-jaw overlap as exposure without
    partial-height weighting; geometric metrics use the actual exposed height.
    """
    ds = pydicom.dcmread(rtplan_path, force=True)
    if getattr(ds, "Modality", None) != "RTPLAN":
        raise ValueError(f"Not an RTPLAN: {rtplan_path}")

    arccheck = _arccheck_from_y_jaws(ds)
    mcs_matlab = _calculate_mcs_matlab(ds)
    alpo_matlab = _calculate_alpo_matlab(ds)

    fg = ds.FractionGroupSequence[0]
    beam_mu = {
        int(rb.ReferencedBeamNumber): float(rb.BeamMeterset)
        for rb in fg.ReferencedBeamSequence
    }

    beam_results = []  # per-beam metric dicts + MU
    total_LT = 0.0
    total_LT_mu = 0.0  # MU weighting for LT

    for beam in ds.BeamSequence:
        if beam.BeamType not in ("DYNAMIC", "STATIC"):
            continue
        mlc_dev = next(
            (d for d in beam.BeamLimitingDeviceSequence
             if d.RTBeamLimitingDeviceType in ("MLCX", "MLCY")),
            None,
        )
        if mlc_dev is None:
            continue
        mlc_type = str(mlc_dev.RTBeamLimitingDeviceType).upper()
        boundaries = np.array([float(x) for x in mlc_dev.LeafPositionBoundaries])
        L = int(mlc_dev.NumberOfLeafJawPairs)
        h_full = np.diff(boundaries)  # nominal leaf widths

        # Walk CPs, carrying forward MLC + jaw positions
        cps = []
        positions = {}
        for cp in beam.ControlPointSequence:
            positions = _device_positions(cp, positions)
            mlc = positions.get(mlc_type)
            if mlc is None:
                continue
            arr = np.asarray(mlc, dtype=float)
            xA_raw, xB_raw = arr[:L], arr[L:]
            jaws = _jaw_extent(positions, mlc_type)
            xA, xB, h_eff = _clip_aperture(xA_raw, xB_raw, boundaries, jaws)
            gap = np.maximum(0.0, xB - xA)
            A = float(np.sum(h_eff * gap))
            P = _aperture_perimeter(xA, xB, h_eff)
            # LSV: only leaves with Y-jaw exposure ("field-defining"),
            # using raw (un-collapsed) leaf positions so closed leaves don't bias.
            active = h_eff > 0
            if np.any(active):
                lsv = _lsv_bank(xA_raw[active]) * _lsv_bank(xB_raw[active])
            else:
                lsv = 1.0
            cps.append({
                "cum_w": float(cp.CumulativeMetersetWeight),
                "xA_raw": xA_raw, "xB_raw": xB_raw,  # for LT
                "xA": xA, "xB": xB,
                "gap": gap, "h_eff": h_eff,
                "A": A, "P": P,
                "LSV": lsv,
            })

        if len(cps) < 2:
            continue

        final_w = float(beam.FinalCumulativeMetersetWeight) or 1.0
        bm = beam_mu.get(int(beam.BeamNumber), 0.0)

        # Per-beam AAV (Masi 2013 / McNiven 2010): each CP's summed aperture (gap)
        # over the sum of every leaf pair's own maximum aperture across the arc.
        max_gap = np.zeros(L)
        for cp in cps:
            max_gap = np.maximum(max_gap, cp["gap"])
        aav_denom = float(np.sum(max_gap)) or 1.0
        for cp in cps:
            cp["AAV"] = float(np.sum(cp["gap"])) / aav_denom

        # CP weights = half the adjacent MU intervals (fraction of beam MU)
        intervals = [
            (cps[i + 1]["cum_w"] - cps[i]["cum_w"]) / final_w
            for i in range(len(cps) - 1)
        ]
        N = len(cps)
        w = np.zeros(N)
        for i in range(N):
            left = intervals[i - 1] if i > 0 else 0.0
            right = intervals[i] if i < N - 1 else 0.0
            w[i] = 0.5 * (left + right)
        wsum = float(np.sum(w)) or 1.0

        # Accumulate per-beam (CP-weighted)
        AAV_b = LSV_b = 0.0
        PA_b = PI_b = 0.0
        EM_num = EM_den = 0.0
        AFW_num = AFW_den = 0.0
        SAS_n = {2: 0.0, 5: 0.0, 10: 0.0, 20: 0.0}
        SAS_w = 0.0
        for i, cp in enumerate(cps):
            wi = w[i]
            if wi <= 0:
                continue
            A, P, gap, h_eff = cp["A"], cp["P"], cp["gap"], cp["h_eff"]
            open_mask = gap > 0
            n_open = int(np.count_nonzero(open_mask))
            h_open_sum = float(np.sum(h_eff[open_mask]))

            AAV_b += wi * cp["AAV"]
            LSV_b += wi * cp["LSV"]
            PA_b  += wi * A
            if A > 0:
                # Edge Metric (Younge 2012, C1=0): leaf-side edges only = full
                # perimeter minus the two leaf-tip edges (2*sum of open leaf heights).
                y_side = P - 2.0 * h_open_sum
                EM_num += wi * (y_side / A)
                EM_den += wi
                PI_b  += wi * (P * P) / (4.0 * math.pi * A)
            if h_open_sum > 0:
                AFW_num += wi * float(np.sum(h_eff[open_mask] * gap[open_mask])) / h_open_sum
                AFW_den += wi
            if n_open > 0:
                # SAS (Crowe 2014): MU-weighted mean of per-CP leaf-pair count ratios
                # N(0 < gap < t) / N(gap > 0).
                SAS_w += wi
                for t in (2, 5, 10, 20):
                    n_small = int(np.count_nonzero((gap > 0) & (gap < t)))
                    SAS_n[t] += wi * (n_small / n_open)

        # MCSv (interval-trapezoidal) and Leaf Travel (raw mechanical positions).
        MCSv_b = 0.0
        for i, dw in enumerate(intervals):
            avgLSV = 0.5 * (cps[i]["LSV"] + cps[i + 1]["LSV"])
            avgAAV = 0.5 * (cps[i]["AAV"] + cps[i + 1]["AAV"])
            MCSv_b += dw * avgLSV * avgAAV
        LT_b = _mean_leaf_travel(cps)

        beam_results.append({
            "mu": bm,
            "MCSv": MCSv_b,
            "AAV":  AAV_b / wsum,
            "LSV":  LSV_b / wsum,
            "AFW":  (AFW_num / AFW_den) if AFW_den > 0 else None,
            "SAS2":  (SAS_n[2]  / SAS_w) if SAS_w > 0 else None,
            "SAS5":  (SAS_n[5]  / SAS_w) if SAS_w > 0 else None,
            "SAS10": (SAS_n[10] / SAS_w) if SAS_w > 0 else None,
            "SAS20": (SAS_n[20] / SAS_w) if SAS_w > 0 else None,
            "PA":   PA_b / wsum,
            "EM":   (EM_num / EM_den) if EM_den > 0 else None,
            "PI":   PI_b / wsum,   # Du 2014 Plan Irregularity (MU-weighted P²/4πA)
            "LT":   LT_b,
        })

    if not beam_results:
        metrics = {}
        if arccheck is not None:
            metrics["ArcCheck"] = arccheck
        if mcs_matlab is not None:
            metrics["MCS_Matlab"] = mcs_matlab
        if alpo_matlab is not None:
            metrics["ALPO_Matlab"] = alpo_matlab
        return _round_metrics(metrics)

    total_mu = sum(b["mu"] for b in beam_results) or 1.0

    def plan_avg(key):
        vals = [(b[key], b["mu"]) for b in beam_results if b[key] is not None]
        if not vals:
            return None
        s = sum(v * mu for v, mu in vals)
        wsum = sum(mu for _, mu in vals) or 1.0
        return s / wsum

    metrics = {
        "ArcCheck": arccheck,
        "TotalMU": total_mu,
        "MCSv": plan_avg("MCSv"),
        "MCS_Matlab": mcs_matlab,
        "ALPO_Matlab": alpo_matlab,
        "AAV":  plan_avg("AAV"),
        "LSV":  plan_avg("LSV"),
        "AFW":  plan_avg("AFW"),
        "SAS2":  plan_avg("SAS2"),
        "SAS5":  plan_avg("SAS5"),
        "SAS10": plan_avg("SAS10"),
        "SAS20": plan_avg("SAS20"),
        "PA":   plan_avg("PA"),
        "EM":   plan_avg("EM"),
        "PI":   plan_avg("PI"),
        "LT":   plan_avg("LT"),
    }
    # Masi 2013 leaf-travel indices, derived from plan-level LT (mm) and MCSv.
    lt_plan = metrics.get("LT")
    if lt_plan is not None:
        metrics["LTi"] = (1000.0 - lt_plan) / 1000.0
        mcsv_plan = metrics.get("MCSv")
        if mcsv_plan is not None:
            metrics["LTMCS"] = metrics["LTi"] * mcsv_plan
    return _round_metrics(metrics)


def find_tps_plan(dcm_files: list[Path], patient_id: str, calc_name: str) -> Path | None:
    """Find the selected DICOM RTPLAN matching {PatientID}_{CalcName}*.dcm."""
    if not patient_id or not calc_name:
        return None

    plan_token = _extract_plan_token(calc_name) or str(calc_name).strip()
    if not plan_token:
        return None

    prefix = f"{patient_id}_{plan_token}".upper()
    candidates = [
        p for p in dcm_files
        if (
            p.suffix.lower() == ".dcm"
            and p.name.upper().startswith(prefix)
            and not p.name.lower().endswith("_dose.dcm")
        )
    ]
    return candidates[0] if candidates else None


def is_structure_set_file(path: Path) -> bool:
    """Return True for RT Structure Set DICOM files by local filename convention."""
    return path.suffix.lower() == ".dcm" and "strctrsets" in path.name.lower()


def find_rt_structure_set(rtstruct_files: list[Path], patient_id: str) -> Path | None:
    """Find the selected RT Structure Set matching {PatientID}_StrctrSets*.dcm."""
    if not patient_id:
        return None

    prefix = f"{patient_id}_".upper()
    candidates = [
        p for p in rtstruct_files
        if is_structure_set_file(p) and p.name.upper().startswith(prefix)
    ]
    return candidates[0] if candidates else None


def _rx_value(rx_gy) -> float | None:
    """Return a numeric prescription dose in Gy, if available."""
    if rx_gy is None:
        return None
    try:
        value = float(rx_gy)
    except (TypeError, ValueError):
        return None
    return value if math.isfinite(value) else None


def _target_roi_priority(roi_name: str, target: str, rx_gy) -> tuple[int, float] | None:
    """Rank target ROI names from most to least specific."""
    name = str(roi_name or "").strip()
    if not name:
        return None

    target = target.upper()
    compact = re.sub(r"\s+", "", name).upper()
    dose_match = re.fullmatch(rf"{re.escape(target)}[_-]?(\d+(?:\.\d+)?)(?:GY)?", compact)
    rx_value = _rx_value(rx_gy)
    if dose_match and rx_value is not None:
        roi_dose = float(dose_match.group(1))
        if math.isclose(roi_dose, rx_value, abs_tol=1e-6):
            return (0, 0.0)

    if dose_match:
        roi_dose = float(dose_match.group(1))
        return (1, -roi_dose)

    if compact == target:
        return (2, 0.0)

    if compact.startswith(target):
        return (3, 0.0)

    return None


def _polygon_area_mm2(points: np.ndarray) -> float:
    """Area of one closed planar contour projected onto the axial XY plane."""
    if len(points) < 3:
        return 0.0
    x = points[:, 0]
    y = points[:, 1]
    return 0.5 * abs(
        float(np.dot(x, np.roll(y, -1)) - np.dot(y, np.roll(x, -1)))
    )


def _roi_volume_cm3(roi_contour) -> float | None:
    """Calculate an RTSTRUCT ROI volume from axial closed planar contours."""
    areas_by_z: dict[float, float] = {}
    for contour in getattr(roi_contour, "ContourSequence", []):
        contour_type = str(getattr(contour, "ContourGeometricType", "CLOSED_PLANAR")).upper()
        if contour_type != "CLOSED_PLANAR":
            continue

        contour_data = getattr(contour, "ContourData", None)
        if not contour_data:
            continue

        points = np.asarray(contour_data, dtype=float).reshape(-1, 3)
        area = _polygon_area_mm2(points)
        if area <= 0:
            continue

        z_key = round(float(np.mean(points[:, 2])), 3)
        areas_by_z[z_key] = areas_by_z.get(z_key, 0.0) + area

    if not areas_by_z:
        return None

    slices = sorted(areas_by_z.items())
    if len(slices) < 2:
        return None

    # Slab method: each contour represents one CT slice, so the volume is the
    # sum of (slice area x slice thickness) — matching the volume a TPS reports.
    # The thickness is the median of inter-slice z-spacings (the nominal CT slice
    # thickness, robust to occasional missing slices); genuine gaps are left
    # un-bridged because absent slices simply contribute no area.
    z_values = [z for z, _ in slices]
    slice_thickness = float(np.median(np.diff(z_values)))
    total_area_mm2 = sum(area for _, area in slices)
    volume_mm3 = total_area_mm2 * slice_thickness

    return round(volume_mm3 / 1000.0, 3)


def _calculate_target_volume_cm3(
    rtstruct_path: str,
    target: str,
    rx_gy=None,
    warning_messages: list[str] | None = None,
) -> object:
    """Calculate target ROI volume in cm3 from an RT Structure Set DICOM."""
    path = Path(rtstruct_path)
    ds = pydicom.dcmread(str(path), force=True, stop_before_pixels=True)
    if getattr(ds, "Modality", None) != "RTSTRUCT":
        raise ValueError(f"Not an RT Structure Set: {rtstruct_path}")

    target = target.upper()
    roi_names = {
        int(roi.ROINumber): str(getattr(roi, "ROIName", "")).strip()
        for roi in getattr(ds, "StructureSetROISequence", [])
        if getattr(roi, "ROINumber", None) is not None
    }
    contours_by_roi = {
        int(contour.ReferencedROINumber): contour
        for contour in getattr(ds, "ROIContourSequence", [])
        if getattr(contour, "ReferencedROINumber", None) is not None
    }

    candidates = []
    for order, (roi_number, roi_name) in enumerate(roi_names.items()):
        priority = _target_roi_priority(roi_name, target, rx_gy)
        if priority is None:
            continue
        roi_contour = contours_by_roi.get(roi_number)
        if roi_contour is None:
            continue
        candidates.append((priority, order, roi_number, roi_name, roi_contour))

    if not candidates:
        return None

    candidates.sort(key=lambda item: (item[0], item[1]))
    selected = candidates[0]
    if len(candidates) > 1:
        candidate_names = ", ".join(item[3] for item in candidates)
        warning = (
            f"Warning: {path.name} has multiple {target} ROI candidates: "
            f"{candidate_names}; using {selected[3]}"
        )
        if warning_messages is not None:
            warning_messages.append(warning)
        print(warning)

    return _roi_volume_cm3(selected[4])


def calculate_ptv(
    rtstruct_path: str,
    rx_gy=None,
    warning_messages: list[str] | None = None,
) -> object:
    """Calculate PTV volume in cm3 from an RT Structure Set DICOM."""
    return _calculate_target_volume_cm3(rtstruct_path, "PTV", rx_gy, warning_messages)


def calculate_ctv(
    rtstruct_path: str,
    rx_gy=None,
    warning_messages: list[str] | None = None,
) -> object:
    """Calculate CTV volume in cm3 from an RT Structure Set DICOM."""
    return _calculate_target_volume_cm3(rtstruct_path, "CTV", rx_gy, warning_messages)


# ── Main ──
COLUMNS = [
    "ID", "Date", "Plan", "Site", "Energy", "ArcCheck", "ModFactor",
    "ALPO(Matlab)", "ALPO(RadCalc)", "Diff(RadCalc)", "AvgFieldSize",
    "AvgDep", "AvgEquPathLength", "CP", "Rx", "Fx", "Dose",
    "GPR3mm", "GPR2mm", "GPR1mm", "GPR3p2mm", "MU", "Gan",
    "Pass_Fail", "Pass",
    "PMU", "AAV", "LSV", "MCSv(Python)", "MCS(Matlab)", "LT", "LTi", "LTMCS",
    "AFW", "SAS2", "SAS5", "SAS10", "SAS20",
    "PA_cm2", "EM", "PI", "PTV", "CTV",
]


def _anonymize_id(pid) -> str | None:
    """Shift each letter to the next letter (Z->A, z->a) and each digit +1 mod 10."""
    if pid is None:
        return None
    out = []
    for ch in str(pid):
        if "A" <= ch <= "Z":
            out.append("A" if ch == "Z" else chr(ord(ch) + 1))
        elif "a" <= ch <= "z":
            out.append("a" if ch == "z" else chr(ord(ch) + 1))
        elif ch.isdigit():
            out.append(str((int(ch) + 1) % 10))
        else:
            out.append(ch)
    return "".join(out)


def _record_to_row_map(rec: dict) -> dict:
    """Convert one extracted PDF record into a spreadsheet row map."""
    b1 = rec.get("block1", {})
    b2 = rec.get("block2", {})
    b3 = rec.get("block3", {})
    b4 = rec.get("block4", {})
    m = rec.get("metrics", {}) or {}
    gpr = rec.get("gpr", {}) or {}

    calc_name = b1.get("Calculation Name", "")
    patient_id = b1.get("Patient ID")
    output_id = patient_id if rec.get("_already_deidentified") else _anonymize_id(patient_id)
    plan_mu = b3.get("Plan MU")
    dose_fx = b2.get("Dose Per Treatment (Gy)")
    pmu = (                                  # Du 2014 PMU normalized to a 2 Gy fraction
        plan_mu * 2.0 / dose_fx
        if isinstance(plan_mu, (int, float)) and isinstance(dose_fx, (int, float)) and dose_fx
        else None
    )

    return {
        "ID":               output_id,
        "Date":             _extract_date(b4.get("Calculated By", "")),
        "Plan":             calc_name,
        "Site":             _extract_site(calc_name),
        "Energy":           b3.get("Energy"),
        "ArcCheck":         m.get("ArcCheck"),
        "ModFactor":        b3.get("Mod Factor"),
        "ALPO(Matlab)":     m.get("ALPO_Matlab"),
        "ALPO(RadCalc)":    b3.get("ALPO"),
        "Diff(RadCalc)":    _num_only(b3.get("Percent Diff")),
        "AvgFieldSize":     b3.get("Eq. Sq."),
        "AvgDep":           b3.get("Avg. Depth"),
        "AvgEquPathLength": b3.get("Avg. Eff. Depth"),
        "CP":               b3.get("Number of Control Points"),
        "Rx":               b2.get("Prescribed Dose (Gy)"),
        "Fx":               b2.get("Num Fractions"),
        "Dose":             b2.get("Dose Per Treatment (Gy)"),
        "GPR3mm":           gpr.get("GPR3mm"),
        "GPR2mm":           gpr.get("GPR2mm"),
        "GPR1mm":           gpr.get("GPR1mm"),
        "GPR3p2mm":         gpr.get("GPR3p2mm"),
        "MU":               plan_mu,
        "Gan":              m.get("GAN"),
        "PMU":              pmu,
        "AAV":              m.get("AAV"),
        "LSV":              m.get("LSV"),
        "MCSv(Python)":     m.get("MCSv"),
        "MCS(Matlab)":      m.get("MCS_Matlab"),
        "LT":               m.get("LT"),
        "LTi":              m.get("LTi"),
        "LTMCS":            m.get("LTMCS"),
        "AFW":              (m.get("AFW") / 10.0) if isinstance(m.get("AFW"), (int, float)) else None,
        "SAS2":             m.get("SAS2"),
        "SAS5":             m.get("SAS5"),
        "SAS10":            m.get("SAS10"),
        "SAS20":            m.get("SAS20"),
        "PA_cm2":           (m.get("PA") / 100.0) if isinstance(m.get("PA"), (int, float)) else None,
        "EM":               m.get("EM"),
        "PI":               m.get("PI"),
        "PTV":              rec.get("PTV"),
        "CTV":              rec.get("CTV"),
    }


def _row_key(row: dict) -> tuple:
    return (row.get("ID"), row.get("Date"), row.get("Site"))


def _load_existing_rows(output_path: str) -> list[dict]:
    if not os.path.exists(output_path):
        return []

    wb = load_workbook(output_path, data_only=True)
    ws = wb.active
    headers = [cell.value for cell in ws[1]]
    rows = []
    for values in ws.iter_rows(min_row=2, values_only=True):
        row_map = {col: None for col in COLUMNS}
        for idx, header in enumerate(headers):
            if header in row_map and idx < len(values):
                row_map[header] = values[idx]
        if any(value is not None for value in row_map.values()):
            rows.append(row_map)
    return rows


def save_to_xlsx(records: list[dict], output_path: str) -> list[str]:
    """Accumulate extracted records into xlsx, replacing matching ID/Date/Site rows."""
    existing_rows = _load_existing_rows(output_path)
    row_index = {_row_key(row): idx for idx, row in enumerate(existing_rows)}
    changes = []

    for rec in records:
        row_map = _record_to_row_map(rec)
        key = _row_key(row_map)
        key_text = " | ".join(str(part or "") for part in key)
        if key in row_index:
            existing_rows[row_index[key]] = row_map
            changes.append(f"Modified: {key_text}")
        else:
            row_index[key] = len(existing_rows)
            existing_rows.append(row_map)
            changes.append(f"Added: {key_text}")

    wb = Workbook()
    ws = wb.active
    ws.title = "RadCalc"

    header_font = Font(name="Arial", bold=True, size=10, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="4472C4")
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin_border = Border(bottom=Side(style="thin", color="B0B0B0"))
    data_font = Font(name="Arial", size=10)
    data_align = Alignment(horizontal="center", vertical="center")

    for col_idx, col_name in enumerate(COLUMNS, start=1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align

    for row_idx, row_map in enumerate(existing_rows, start=2):
        for col_idx, col_name in enumerate(COLUMNS, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=row_map.get(col_name))
            cell.font = data_font
            cell.alignment = data_align
            cell.border = thin_border

    for col_idx, col_name in enumerate(COLUMNS, start=1):
        ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = max(
            len(col_name) + 4, 10
        )

    ws.freeze_panes = "A2"
    wb.save(output_path)
    return changes


def _select_files(
    title: str,
    filetypes: list[tuple[str, str]],
    initial_dir: Path | None = None,
) -> list[Path]:
    from tkinter import Tk, filedialog

    root = Tk()
    root.withdraw()
    root.attributes("-topmost", True)
    dialog_options = {"title": title, "filetypes": filetypes}
    if initial_dir is not None:
        existing_initial_dir = _existing_initial_dir(initial_dir)
        if existing_initial_dir:
            dialog_options["initialdir"] = existing_initial_dir
    paths = filedialog.askopenfilenames(**dialog_options)
    root.destroy()
    return [Path(p) for p in paths]


def _print_intro() -> None:
    print("=" * 64)
    print("  Data Extract  -  RadCalc QA Tool")
    print("=" * 64)
    print()
    print("  This tool reads RadCalc PDF reports together with the TPS")
    print("  RTPLAN DICOM files, and writes the")
    print("  combined QA data into 'data.xlsx' next to this program.")
    print()
    print("  You will be asked to select three groups of files in order:")
    print("    1. RadCalc PDF reports   (*.pdf)")
    print("    2. TPS RTPLAN DICOM      (*.dcm)   - NOT the *_dose.dcm")
    print("    3. RT Structure Set DICOM (*.dcm)")
    print()
    print("  Notes:")
    print("    - Files may be located anywhere on disk.")
    print("    - PDFs drive the row set; DCM files are matched into PDF")
    print("      rows and never produce rows on their own.")
    print("    - Patient ID must be unique across the selected PDF rows.")
    print("    - 'data.xlsx' is cumulative. Rows with the same")
    print("      ID + Date + Site are updated in place; new rows are")
    print("      appended.")
    print("    - If a DCM match is missing, the PDF row is still written")
    print("      but the related metrics / Gan columns will be left blank,")
    print("      and an 'Error:' message will be shown.")
    print()
    print("  Please make sure the PDF and DCM belong to the same patient")
    print("  and the same plan.")
    print()
    print("=" * 64)
    print()


def main() -> None:
    _print_intro()
    if getattr(sys, "frozen", False):
        script_dir = os.path.dirname(sys.executable)
    else:
        script_dir = os.path.dirname(os.path.abspath(__file__))
    history_path = Path(script_dir) / HISTORY_FILE_NAME
    history = _load_history(history_path)
    pdf_initial_dir = _history_initial_dir(history["pdf_dirs"], PDF_INITIAL_DIR)
    dcm_initial_dir = _history_initial_dir(history["dcm_dirs"], DCM_INITIAL_DIR)
    rtstruct_initial_dir = _history_initial_dir(
        history["rtstruct_dirs"],
        RTSTRUCT_INITIAL_DIR,
    )

    pdf_files = _select_files(
        "Select RadCalc PDF file(s)",
        [("PDF files", "*.pdf"), ("All files", "*.*")],
        pdf_initial_dir,
    )
    if not pdf_files:
        message = "Error: no PDF files selected"
        print(message)
        _prepend_history_messages(history, [message])
        _save_history(history_path, history)
        raise SystemExit(1)
    _record_history_dir(history, "pdf_dirs", pdf_files)

    dcm_files = _select_files(
        "Select TPS DICOM plan file(s) (*.dcm)",
        [("DICOM files", "*.dcm"), ("All files", "*.*")],
        dcm_initial_dir,
    )
    ignored_structure_set_files = [path for path in dcm_files if is_structure_set_file(path)]
    dcm_files = [path for path in dcm_files if not is_structure_set_file(path)]
    _record_history_dir(history, "dcm_dirs", dcm_files)

    rtstruct_files = _select_files(
        "Select RT Structure Set DICOM file(s)",
        [("DICOM files", "*.dcm"), ("All files", "*.*")],
        rtstruct_initial_dir,
    )
    ignored_plan_files = [path for path in rtstruct_files if not is_structure_set_file(path)]
    rtstruct_files = [path for path in rtstruct_files if is_structure_set_file(path)]
    _record_history_dir(history, "rtstruct_dirs", rtstruct_files)

    errors = []
    errors.extend(
        f"Error: {path.name} ignored because RT Structure Set DICOM files must be selected in the third window"
        for path in ignored_structure_set_files
    )
    errors.extend(
        f"Error: {path.name} ignored because TPS DICOM plan files must be selected in the second window"
        for path in ignored_plan_files
    )
    if len(dcm_files) != len(pdf_files):
        errors.append(
            f"Error: selected {len(dcm_files)} DCM file(s), expected {len(pdf_files)}"
        )

    all_params = []
    warning_messages = []
    matched_rtstruct_files: set[Path] = set()
    for pdf_path in pdf_files:
        try:
            params = extract_radcalc_params(str(pdf_path))
        except Exception as e:
            errors.append(f"Error: {pdf_path.name} | failed to read PDF | {e}")
            continue

        warning_messages.extend(params.get("_warnings", []))

        pid = params["block1"].get("Patient ID")
        calc = params["block1"].get("Calculation Name")
        params["metrics"] = {}
        params["PTV"] = None
        params["CTV"] = None

        tps_plan = find_tps_plan(dcm_files, pid, calc)
        if tps_plan is None:
            output_id = pid if params.get("_already_deidentified") else _anonymize_id(pid)
            errors.append(
                f"Error: {output_id} | {pid} | {calc} | matching DCM not found"
            )
        else:
            try:
                params["metrics"].update(extract_complexity_metrics(str(tps_plan)))
                gan = extract_gan_from_plan(str(tps_plan))
                if gan is not None:
                    params["metrics"]["GAN"] = gan
            except Exception as e:
                errors.append(f"Error: {pid} | {calc} | {tps_plan.name} | {e}")

        rtstruct = find_rt_structure_set(rtstruct_files, pid)
        if rtstruct is None:
            output_id = pid if params.get("_already_deidentified") else _anonymize_id(pid)
            errors.append(
                f"Error: {output_id} | {pid} | {calc} | matching RT Structure Set not found"
            )
        else:
            matched_rtstruct_files.add(rtstruct)
            try:
                params["PTV"] = calculate_ptv(
                    str(rtstruct),
                    params["block2"].get("Prescribed Dose (Gy)"),
                    warning_messages,
                )
                params["CTV"] = calculate_ctv(
                    str(rtstruct),
                    params["block2"].get("Prescribed Dose (Gy)"),
                    warning_messages,
                )
            except Exception as e:
                errors.append(f"Error: {pid} | {calc} | {rtstruct.name} | {e}")

        all_params.append(params)

    errors.extend(
        f"Error: {path.name} selected in the third window but did not match any PDF row"
        for path in rtstruct_files
        if path not in matched_rtstruct_files
    )

    out_path = os.path.join(script_dir, "data.xlsx")
    changes = save_to_xlsx(all_params, out_path) if all_params else []
    for message in changes + errors:
        print(message)
    _prepend_history_messages(history, changes + warning_messages + errors)
    _save_history(history_path, history)

    if getattr(sys, "frozen", False):
        input("\nPress Enter to exit...")


if __name__ == "__main__":
    main()
